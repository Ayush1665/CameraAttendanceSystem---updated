import tkinter as tk
from tkinter import ttk
import cv2
import os
import pyttsx3
import openpyxl

def add_user():
    cap = cv2.VideoCapture(0)
    root = tk.Tk()
    root.title("Add User")
    root.geometry("400x300")

    # Apply theme
    style = ttk.Style(root)
    root.tk.call("source", "azure.tcl")
    style.theme_use("azure")

    engine = pyttsx3.init()

    def capture_image():
        name = name_entry.get()
        roll_no = roll_entry.get()

        if name.strip() == "" or roll_no.strip() == "":
            status_label.config(text="Name and Roll No. cannot be empty", fg="red")
            return

        if os.path.exists('users.xlsx'):
            workbook = openpyxl.load_workbook('users.xlsx')
            sheet = workbook["Users"]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                existing_roll_no, existing_name = row
                if roll_no == existing_roll_no:
                    engine.say("User already exists")
                    engine.runAndWait()
                    status_label.config(text="User already exists", fg="red")
                    return
        else:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Users"
            sheet.append(["Roll No", "Name"])

        while True:
            ret, frame = cap.read()
            if not ret:
                status_label.config(text="Failed to capture image", fg="red")
                return

            cv2.imshow("Capture Image", frame)
            if cv2.waitKey(1) & 0xFF == ord('c'):
                break

        file_path = os.path.join("users", f"{name}_{roll_no}.jpg")
        if not os.path.exists("users"):
            os.makedirs("users")
        cv2.imwrite(file_path, frame)
        
        sheet.append([roll_no, name])
        workbook.save('users.xlsx')

        engine.say("User added")
        engine.runAndWait()
        status_label.config(text="User added", fg="green")

        cv2.destroyAllWindows()
        cap.release()
        root.destroy()

    name_label = ttk.Label(root, text="Name:")
    name_label.pack(pady=5)
    name_entry = ttk.Entry(root)
    name_entry.pack(pady=5)

    roll_label = ttk.Label(root, text="Roll No.:")
    roll_label.pack(pady=5)
    roll_entry = ttk.Entry(root)
    roll_entry.pack(pady=5)

    capture_button = ttk.Button(root, text="Capture", command=capture_image)
    capture_button.pack(pady=10)

    status_label = tk.Label(root, text="", font=("Arial", 10))
    status_label.pack()

    root.mainloop()
    cap.release()

if __name__ == "__main__":
    add_user()
