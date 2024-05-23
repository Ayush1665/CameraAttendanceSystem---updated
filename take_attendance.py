import cv2
import os
import pyttsx3
import openpyxl
from recognize import load_known_faces, recognize_faces
from datetime import datetime

def take_attendance():
    cap = cv2.VideoCapture(0)

    if not cap.isOpened():
        print("Error: Could not open camera. Please check if the camera is connected and not being used by another application.")
        return

    engine = pyttsx3.init()

    known_faces, known_names, known_roll_nos = load_known_faces()
    
    spoken_thank_you = False
    spoken_unauthorized = False

    if not os.path.exists('attendance.xlsx'):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Attendance"
        sheet.append(["Roll No", "Name", "Status", "Timestamp"])
    else:
        workbook = openpyxl.load_workbook('attendance.xlsx')
        sheet = workbook["Attendance"]

    existing_roll_nos = {row[0].value for row in sheet.iter_rows(min_row=2, values_only=True)}

    while True:
        ret, frame = cap.read()
        if not ret:
            print("Error: Failed to capture image")
            break

        names, roll_nos = recognize_faces(frame, known_faces, known_names, known_roll_nos)

        if names:
            for name, roll_no in zip(names, roll_nos):
                current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                if roll_no not in existing_roll_nos:
                    sheet.append([roll_no, name, "Present", current_time])
                    existing_roll_nos.add(roll_no)
                else:
                    for row in sheet.iter_rows(min_row=2):
                        if row[0].value == roll_no:
                            row[2].value = "Present"
                            row[3].value = current_time
                            break

                cv2.putText(frame, f"{name} ({roll_no})", (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)

            if not spoken_thank_you:
                engine.say("Thank you")
                engine.runAndWait()
                spoken_thank_you = True
                spoken_unauthorized = False  # Reset unauthorized flag
        else:
            if not spoken_unauthorized:
                engine.say("Unauthorized User")
                engine.runAndWait()
                spoken_unauthorized = True
                spoken_thank_you = False  # Reset thank you flag

        cv2.imshow("Attendance", frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    workbook.save('attendance.xlsx')
    cap.release()
    cv2.destroyAllWindows()

if __name__ == "__main__":
    take_attendance()
